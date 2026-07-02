#!/usr/bin/env python3
"""Generate a task quality-check Excel template from examples_per_task."""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_DIR = Path(__file__).resolve().parents[1]
DEFAULT_TASKS_DIR = REPO_DIR / "evaluation_examples" / "examples_per_task"
DEFAULT_OUTPUT = REPO_DIR / "task_qc.xlsx"

HEADERS = ["任务id", "质检人", "质检日期", "是否可用", "错误类型", "修复方式"]
USABLE_OPTIONS = ["是", "否", "待检"]
ERROR_TYPE_OPTIONS = ["evaluator", "setup_file", "指令可行性", "其他"]
COLUMN_WIDTHS = {
    "A": 60,
    "B": 12,
    "C": 14,
    "D": 10,
    "E": 16,
    "F": 40,
}


def collect_task_ids(tasks_dir: Path) -> list[str]:
    task_ids: list[str] = []
    for domain_dir in sorted(tasks_dir.iterdir()):
        if not domain_dir.is_dir():
            continue
        for task_dir in sorted(domain_dir.iterdir()):
            if task_dir.is_dir() and (task_dir / "task.json").is_file():
                task_ids.append(f"{domain_dir.name}/{task_dir.name}")
    return task_ids


def add_list_validation(ws, col_letter: str, options: list[str], max_row: int) -> None:
    formula = f'"{",".join(options)}"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def build_workbook(task_ids: list[str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "任务质检"

    ws.append(HEADERS)
    for task_id in task_ids:
        ws.append([task_id, "", "", "", "", ""])

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(len(task_ids) + 1, 1)}"

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    max_row = len(task_ids) + 1
    if max_row >= 2:
        add_list_validation(ws, "D", USABLE_OPTIONS, max_row)
        add_list_validation(ws, "E", ERROR_TYPE_OPTIONS, max_row)

    return wb


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate task QC Excel template.")
    parser.add_argument(
        "--tasks-dir",
        type=Path,
        default=DEFAULT_TASKS_DIR,
        help="Path to examples_per_task directory.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Output .xlsx path.",
    )
    args = parser.parse_args()

    tasks_dir = args.tasks_dir.resolve()
    if not tasks_dir.is_dir():
        raise SystemExit(f"Tasks directory not found: {tasks_dir}")

    task_ids = collect_task_ids(tasks_dir)
    if not task_ids:
        raise SystemExit(f"No tasks found under: {tasks_dir}")

    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook(task_ids)
    wb.save(output_path)
    print(f"Wrote {len(task_ids)} tasks to {output_path}")


if __name__ == "__main__":
    main()
