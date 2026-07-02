#!/usr/bin/env python3
"""One-off fix for calc tasks with polluted/missing initial xlsx files."""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parents[2]


def backup(path: Path) -> Path:
    bak = path.with_suffix(path.suffix + ".bak")
    shutil.copy2(path, bak)
    return bak


TARGET_STAFF_HEADERS = ["ID", "Name", "Department", "Join Date", "Salary"]
SCRAMBLED_STAFF_HEADERS = ["Name", "Salary", "Department", "Join Date", "ID"]


def fix_gross_margin(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    if all(ws.cell(row=row, column=4).value is None for row in range(2, ws.max_row + 1)):
        wb.close()
        print(f"skip gross margin (already clean): {path}")
        return
    backup(path)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).value = None
    wb.save(path)
    wb.close()
    print(f"fixed gross margin: cleared D2:D{ws.max_row} in {path}")


def fix_chart_workbook(path: Path) -> None:
    """Keep only Sheet1 data; remove charts and extra sheets."""
    wb = openpyxl.load_workbook(path)
    if wb.sheetnames == ["Sheet1"] and not getattr(wb["Sheet1"], "_charts", []):
        wb.close()
        print(f"skip chart workbook (already clean): {path}")
        return
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet1 missing in {path}")
    backup(path)

    # Rebuild workbook with only Sheet1, no charts.
    data_wb = openpyxl.Workbook()
    src = wb["Sheet1"]
    dst = data_wb.active
    dst.title = "Sheet1"
    for row in src.iter_rows():
        for cell in row:
            dst[cell.coordinate].value = cell.value

    data_wb.save(path)
    wb.close()
    print(f"fixed chart workbook: kept Sheet1 only, no charts in {path}")


def fix_emp_stats(path: Path) -> None:
    """Keep only Employees sheet; agent must create Stats."""
    wb = openpyxl.load_workbook(path)
    if wb.sheetnames == ["Employees"]:
        wb.close()
        print(f"skip emp stats (already clean): {path}")
        return
    if "Employees" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Employees sheet missing in {path}")
    backup(path)

    data_wb = openpyxl.Workbook()
    src = wb["Employees"]
    dst = data_wb.active
    dst.title = "Employees"
    for row in src.iter_rows():
        for cell in row:
            dst[cell.coordinate].value = cell.value

    data_wb.save(path)
    wb.close()
    print(f"fixed emp stats: kept Employees only in {path}")


def fix_staff_records_reorder(path: Path) -> None:
    """Scramble column order so headers no longer match the target layout."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if headers[: len(SCRAMBLED_STAFF_HEADERS)] == SCRAMBLED_STAFF_HEADERS:
        wb.close()
        print(f"skip staff records (already scrambled): {path}")
        return
    if headers[: len(TARGET_STAFF_HEADERS)] != TARGET_STAFF_HEADERS:
        wb.close()
        raise ValueError(f"unexpected headers {headers} in {path}")
    backup(path)

    rows: list[dict[str, object | None]] = []
    for row_idx in range(1, ws.max_row + 1):
        rows.append(
            {header: ws.cell(row_idx, col + 1).value for col, header in enumerate(TARGET_STAFF_HEADERS)}
        )

    for row_idx, row_data in enumerate(rows, start=1):
        for col_idx, header in enumerate(SCRAMBLED_STAFF_HEADERS, start=1):
            ws.cell(row_idx, col_idx).value = (
                header if row_idx == 1 else row_data[header]
            )

    wb.save(path)
    wb.close()
    print(f"fixed staff records: scrambled columns to {SCRAMBLED_STAFF_HEADERS} in {path}")


def main() -> None:
    fixes = [
        (
            "8c891b4d gross margin (cache)",
            PROJECT_ROOT / "cache/8c891b4d-2e1e-4a79-9a4d-2e326c6d5ae9/Product_Gross_Margin.xlsx",
            fix_gross_margin,
        ),
        (
            "8c891b4d gross margin (assets)",
            PROJECT_ROOT / "assets/xlsx/Product_Gross_Margin.xlsx",
            fix_gross_margin,
        ),
        (
            "f7c9725b chart (cache)",
            PROJECT_ROOT / "cache/f7c9725b-9f02-4131-a007-68de65161a3c/Region_SalesVsTarget_Chart.xlsx",
            fix_chart_workbook,
        ),
        (
            "f7c9725b chart (assets)",
            PROJECT_ROOT / "assets/xlsx/Region_SalesVsTarget_Chart.xlsx",
            fix_chart_workbook,
        ),
        (
            "4f989467 emp stats",
            PROJECT_ROOT / "assets/xlsx/EmpStats_L2_07.xlsx",
            fix_emp_stats,
        ),
        (
            "fb0f694e staff records (cache, L1)",
            PROJECT_ROOT
            / "cache/fb0f694e-fc10-46f9-9fcd-169630844996/Staff_Records_Reorder.xlsx",
            fix_staff_records_reorder,
        ),
        (
            "fb0f694e staff records (assets, L2)",
            PROJECT_ROOT / "assets/xlsx/Staff_Records_Reorder.xlsx",
            fix_staff_records_reorder,
        ),
    ]

    for label, path, fn in fixes:
        if not path.is_file():
            raise FileNotFoundError(f"{label}: {path}")
        fn(path)

    print(f"All {len(fixes)} initial files processed.")


if __name__ == "__main__":
    main()
