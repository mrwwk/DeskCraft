import logging
import os

logger = logging.getLogger("desktopenv.metric.general")


def check_supplier_audit_xlsx(result_path, expected=None, **options):
    """
    检查 supplier_audit_log.xlsx 第2行是否包含所有期望的审计数据字段。

    Args:
        result_path (str): 从 VM 拉取的 xlsx 文件本地路径。
        expected (dict): 期望的单元格值映射，如 {"A2": "...", "B2": "...", ...}。
        **options: 额外选项（预留）。

    Returns:
        float: 1.0 表示所有字段匹配，0.0 表示不匹配或读取失败。
    """
    if result_path is None or not os.path.exists(result_path):
        logger.warning(f"Result file not found: {result_path}")
        return 0.0

    if expected is None:
        logger.warning("Expected data not provided")
        return 0.0

    try:
        from openpyxl import load_workbook

        wb = load_workbook(result_path, data_only=True)
        ws = wb.active

        for cell_ref, expected_val in expected.items():
            actual_val = ws[cell_ref].value
            # D2 (Overall Score) 可能是数字 90 或字符串 "90"，统一比较
            if isinstance(expected_val, int):
                try:
                    actual_val = int(actual_val) if actual_val is not None else None
                except (ValueError, TypeError):
                    pass
            if actual_val != expected_val:
                logger.warning(
                    f"Cell {cell_ref}: expected '{expected_val}', got '{actual_val}'"
                )
                wb.close()
                return 0.0

        wb.close()
        return 1.0
    except Exception as e:
        logger.error(f"Error reading xlsx file '{result_path}': {e}")
        return 0.0
