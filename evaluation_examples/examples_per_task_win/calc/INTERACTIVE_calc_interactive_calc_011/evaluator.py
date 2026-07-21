"""Evaluator for INTERACTIVE_calc_interactive_calc_011."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_011")


def _row_values(ws, row, max_col):
    return [ws.cell(row=row, column=col).value for col in range(1, max_col + 1)]


def check_first_rows_backup(result_path, expected=None, **options):
    """Verify Sheet2 contains exactly the header plus the first N data rows from Sheet1."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    expected = expected or {}
    source_sheet = expected.get("source_sheet", "Sheet1")
    target_sheet = expected.get("target_sheet", "Sheet2")
    expected_rows = int(expected.get("expected_data_rows", 3))

    try:
        import openpyxl

        wb = openpyxl.load_workbook(result_path, data_only=True)
        if source_sheet not in wb.sheetnames or target_sheet not in wb.sheetnames:
            return 0.0

        ws1 = wb[source_sheet]
        ws2 = wb[target_sheet]
        max_col = ws1.max_column
        expected_content = [_row_values(ws1, row, max_col) for row in range(1, expected_rows + 2)]

        if ws2.max_row != expected_rows + 1 or ws2.max_column != max_col:
            return 0.0

        actual_content = [_row_values(ws2, row, max_col) for row in range(1, expected_rows + 2)]
        return 1.0 if actual_content == expected_content else 0.0
    except Exception as exc:
        logger.warning("check_first_rows_backup failed: %s", exc)
        return 0.0
    finally:
        try:
            wb.close()
        except Exception:
            pass
