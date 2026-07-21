"""Evaluator for INTERACTIVE_calc_interactive_calc_workflow_001."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_workflow_001")


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _total_formula_mentions_months(value, row):
    if not isinstance(value, str) or not value.startswith("="):
        return False
    formula = value.replace("$", "").replace(" ", "").upper()
    if f"B{row}:G{row}" in formula:
        return True
    return all(f"{column}{row}" in formula for column in ["B", "C", "D", "E", "F", "G"])


def check_calc_workflow(result_path, expected=None, **options):
    """Verify Total column, descending sort, Summary sheet, and bold headers."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    expected = expected or {}
    tolerance = float(expected.get("tolerance", 0.01))

    try:
        import openpyxl

        wb_formula = openpyxl.load_workbook(result_path, data_only=False)
        wb_values = openpyxl.load_workbook(result_path, data_only=True)
        if "Sheet1" not in wb_formula.sheetnames or "Summary" not in wb_formula.sheetnames:
            return 0.0

        ws_formula = wb_formula["Sheet1"]
        ws_values = wb_values["Sheet1"]
        summary = wb_values["Summary"]

        headers = [str(ws_formula.cell(row=1, column=col).value).strip() for col in range(1, ws_formula.max_column + 1)]
        if headers[-1].lower() != "total":
            return 0.0
        total_col = len(headers)
        rep_col = 1
        month_cols = list(range(2, total_col))

        totals = []
        reps = []
        for row in range(2, ws_formula.max_row + 1):
            rep = ws_values.cell(row=row, column=rep_col).value
            if rep is None:
                continue
            reps.append(str(rep))
            month_values = []
            for col in month_cols:
                value = _number(ws_values.cell(row=row, column=col).value)
                if value is None:
                    return 0.0
                month_values.append(value)
            expected_total = sum(month_values)
            actual_total = _number(ws_values.cell(row=row, column=total_col).value)
            formula = ws_formula.cell(row=row, column=total_col).value
            if _total_formula_mentions_months(formula, row):
                pass
            elif actual_total is None or abs(actual_total - expected_total) > tolerance:
                return 0.0
            totals.append(actual_total if actual_total is not None else expected_total)

        if any(totals[i] < totals[i + 1] for i in range(len(totals) - 1)):
            return 0.0

        if str(summary["A1"].value).strip() != "Top Performer":
            return 0.0
        if str(summary["B1"].value).strip() != reps[0]:
            return 0.0

        for col in range(1, ws_formula.max_column + 1):
            if not ws_formula.cell(row=1, column=col).font.bold:
                return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_calc_workflow failed: %s", exc)
        return 0.0
    finally:
        for wb in ("wb_formula", "wb_values"):
            try:
                locals()[wb].close()
            except Exception:
                pass
