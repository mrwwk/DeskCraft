"""Evaluator for INTERACTIVE_calc_interactive_calc_005."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_005")


def _key(value):
    return str(value).strip()


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def check_promotion_revenue_summary(result_path, expected=None, **options):
    """Verify Sheet2 summarizes total Revenue by Promotion Type."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    expected = expected or {}
    source_sheet = expected.get("source_sheet", "Sheet1")
    target_sheet = expected.get("target_sheet", "Sheet2")
    group_header = expected.get("group_header", "Promotion Type")
    total_header = expected.get("total_header", "Total Revenue")
    tolerance = float(expected.get("tolerance", 0.01))

    try:
        import openpyxl

        wb = openpyxl.load_workbook(result_path, data_only=True)
        if source_sheet not in wb.sheetnames or target_sheet not in wb.sheetnames:
            return 0.0

        ws1 = wb[source_sheet]
        ws2 = wb[target_sheet]
        source_headers = [str(ws1.cell(row=1, column=col).value).strip() for col in range(1, ws1.max_column + 1)]
        if "Promotion" not in source_headers or "Revenue" not in source_headers:
            return 0.0
        promotion_col = source_headers.index("Promotion") + 1
        revenue_col = source_headers.index("Revenue") + 1

        expected_totals = {}
        for row in range(2, ws1.max_row + 1):
            promotion = ws1.cell(row=row, column=promotion_col).value
            revenue = _number(ws1.cell(row=row, column=revenue_col).value)
            if promotion is None or str(promotion).strip() == "" or revenue is None:
                continue
            expected_totals[_key(promotion)] = expected_totals.get(_key(promotion), 0.0) + revenue

        target_headers = [str(ws2.cell(row=1, column=col).value).strip() for col in range(1, ws2.max_column + 1)]
        if len(target_headers) < 2:
            return 0.0
        if target_headers[0].lower() != group_header.lower() or target_headers[1].lower() != total_header.lower():
            return 0.0

        actual_totals = {}
        for row in range(2, ws2.max_row + 1):
            promotion = ws2.cell(row=row, column=1).value
            total = _number(ws2.cell(row=row, column=2).value)
            if promotion is None or str(promotion).strip() == "":
                continue
            if total is None:
                return 0.0
            actual_totals[_key(promotion)] = total

        if set(actual_totals) != set(expected_totals):
            return 0.0
        for promotion, expected_total in expected_totals.items():
            if abs(actual_totals[promotion] - expected_total) > tolerance:
                return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_promotion_revenue_summary failed: %s", exc)
        return 0.0
    finally:
        try:
            wb.close()
        except Exception:
            pass
