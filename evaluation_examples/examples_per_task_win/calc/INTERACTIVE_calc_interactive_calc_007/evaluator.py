"""Evaluator for INTERACTIVE_calc_interactive_calc_007."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_007")


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _formula_mentions_required_cells(value, row, columns):
    if not isinstance(value, str) or not value.startswith("="):
        return False
    formula = value.replace("$", "").replace(" ", "").upper()
    if not all(f"{column_letter}{row}" in formula for column_letter in columns):
        return False
    return formula.count("-") >= 5


def check_gross_profit(result_path, expected=None, **options):
    """Verify Gross Profit equals sales minus returns, discounts, and all expenses."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    tolerance = float((expected or {}).get("tolerance", 0.01))

    try:
        import openpyxl

        wb_formula = openpyxl.load_workbook(result_path, data_only=False)
        wb_values = openpyxl.load_workbook(result_path, data_only=True)
        ws_formula = wb_formula.active
        ws_values = wb_values[ws_formula.title]

        headers = [str(ws_formula.cell(row=1, column=col).value).strip() for col in range(1, ws_formula.max_column + 1)]
        required = [
            "Sales",
            "Sales Return",
            "Discounts and Allowances",
            "Materials Charges",
            "Labor Charges",
            "Overhead",
            "Gross Profit",
        ]
        if any(header not in headers for header in required):
            return 0.0

        cols = {header: headers.index(header) + 1 for header in required}
        expense_cols = [cols["Sales Return"], cols["Discounts and Allowances"], cols["Materials Charges"], cols["Labor Charges"], cols["Overhead"]]
        required_formula_refs = ["B", "C", "D", "F", "G", "H"]

        for row in range(2, ws_formula.max_row + 1):
            sales = _number(ws_values.cell(row=row, column=cols["Sales"]).value)
            if sales is None:
                continue
            expenses = []
            for col in expense_cols:
                value = _number(ws_values.cell(row=row, column=col).value)
                if value is None:
                    return 0.0
                expenses.append(value)
            expected_value = sales - sum(expenses)

            gross_cell_formula = ws_formula.cell(row=row, column=cols["Gross Profit"]).value
            gross_cell_value = _number(ws_values.cell(row=row, column=cols["Gross Profit"]).value)
            if _formula_mentions_required_cells(gross_cell_formula, row, required_formula_refs):
                continue
            if gross_cell_value is None or abs(gross_cell_value - expected_value) > tolerance:
                return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_gross_profit failed: %s", exc)
        return 0.0
    finally:
        for wb in ("wb_formula", "wb_values"):
            try:
                locals()[wb].close()
            except Exception:
                pass
