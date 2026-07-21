"""Evaluator for INTERACTIVE_calc_interactive_calc_003."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_003")


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _formula_matches(value, row):
    if not isinstance(value, str) or not value.startswith("="):
        return False
    formula = value.replace("$", "").replace(" ", "").upper()
    return f"B{row}" in formula and f"C{row}" in formula and "-" in formula


def check_profit_column(result_path, expected=None, **options):
    """Verify that a rightmost Profit column equals Sales - COGS for every data row."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    tolerance = float((expected or {}).get("tolerance", 0.01))

    try:
        import openpyxl

        wb_formula = openpyxl.load_workbook(result_path, data_only=False)
        wb_values = openpyxl.load_workbook(result_path, data_only=True)
        ws_formula = wb_formula.active
        ws_values = wb_values[ws_formula.title]

        headers = [str(ws_formula.cell(row=1, column=col).value).strip() for col in range(1, ws_formula.max_column + 1)]
        if headers[-1].lower() != "profit":
            return 0.0

        sales_col = headers.index("Sales") + 1 if "Sales" in headers else None
        cogs_col = headers.index("COGS") + 1 if "COGS" in headers else None
        profit_col = len(headers)
        if sales_col is None or cogs_col is None:
            return 0.0

        for row in range(2, ws_formula.max_row + 1):
            sales = _number(ws_values.cell(row=row, column=sales_col).value)
            cogs = _number(ws_values.cell(row=row, column=cogs_col).value)
            if sales is None or cogs is None:
                continue

            expected_value = sales - cogs
            formula_value = ws_formula.cell(row=row, column=profit_col).value
            cached_value = _number(ws_values.cell(row=row, column=profit_col).value)

            if _formula_matches(formula_value, row):
                continue
            if cached_value is not None and abs(cached_value - expected_value) <= tolerance:
                continue
            return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_profit_column failed: %s", exc)
        return 0.0
    finally:
        for wb in ("wb_formula", "wb_values"):
            try:
                locals()[wb].close()
            except Exception:
                pass
