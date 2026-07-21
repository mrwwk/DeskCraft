"""Evaluator for INTERACTIVE_calc_interactive_calc_002."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_002")


def _cell_values(ws, column, start_row, end_row):
    return [ws.cell(row=row, column=column).value for row in range(start_row, end_row + 1)]


def check_revenue_column_copied(result_path, expected=None, **options):
    """Verify that Sheet2 starts at A1 with an exact copy of Sheet1's Revenue column."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    try:
        import openpyxl

        wb = openpyxl.load_workbook(result_path, data_only=True)
        sheet1_name = (expected or {}).get("source_sheet", "Sheet1")
        sheet2_name = (expected or {}).get("target_sheet", "Sheet2")
        header = (expected or {}).get("header", "Revenue")

        if sheet1_name not in wb.sheetnames or sheet2_name not in wb.sheetnames:
            return 0.0

        ws1 = wb[sheet1_name]
        ws2 = wb[sheet2_name]

        revenue_col = None
        for col in range(1, ws1.max_column + 1):
            if str(ws1.cell(row=1, column=col).value).strip().lower() == header.lower():
                revenue_col = col
                break
        if revenue_col is None:
            return 0.0

        source_values = _cell_values(ws1, revenue_col, 1, ws1.max_row)
        target_values = _cell_values(ws2, 1, 1, len(source_values))

        if target_values != source_values:
            logger.warning("Sheet2 column A does not match Sheet1 Revenue column")
            return 0.0

        for col in range(2, ws2.max_column + 1):
            if any(ws2.cell(row=row, column=col).value is not None for row in range(1, ws2.max_row + 1)):
                return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_revenue_column_copied failed: %s", exc)
        return 0.0
    finally:
        try:
            wb.close()
        except Exception:
            pass
