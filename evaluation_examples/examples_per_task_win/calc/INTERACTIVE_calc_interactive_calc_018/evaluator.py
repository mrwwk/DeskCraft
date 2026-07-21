"""Evaluator for INTERACTIVE_calc_interactive_calc_018."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_018")


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def check_revenue_total_on_sheet1(result_path, expected=None, **options):
    """Verify Sheet2 was deleted and the Revenue total is at the bottom of Sheet1."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    expected = expected or {}
    tolerance = float(expected.get("tolerance", 0.01))

    try:
        import openpyxl

        wb = openpyxl.load_workbook(result_path, data_only=True)
        if expected.get("deleted_sheet", "Sheet2") in wb.sheetnames:
            return 0.0
        sheet_name = expected.get("source_sheet", "Sheet1")
        if sheet_name not in wb.sheetnames:
            return 0.0
        ws = wb[sheet_name]

        revenue_col = None
        for col in range(1, ws.max_column + 1):
            if str(ws.cell(row=1, column=col).value).strip().lower() == "revenue":
                revenue_col = col
                break
        if revenue_col is None:
            return 0.0

        numeric_rows = []
        for row in range(2, ws.max_row + 1):
            value = _number(ws.cell(row=row, column=revenue_col).value)
            if value is not None:
                numeric_rows.append((row, value))

        if len(numeric_rows) < 2:
            return 0.0
        total_row, total_value = numeric_rows[-1]
        source_sum = sum(value for _, value in numeric_rows[:-1])
        if total_row <= numeric_rows[-2][0]:
            return 0.0
        if abs(total_value - source_sum) > tolerance:
            return 0.0
        return 1.0
    except Exception as exc:
        logger.warning("check_revenue_total_on_sheet1 failed: %s", exc)
        return 0.0
    finally:
        try:
            wb.close()
        except Exception:
            pass
