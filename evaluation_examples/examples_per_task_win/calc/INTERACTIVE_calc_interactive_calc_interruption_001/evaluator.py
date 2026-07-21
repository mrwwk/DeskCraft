"""Evaluator for INTERACTIVE_calc_interactive_calc_interruption_001."""

import logging
import os

logger = logging.getLogger("desktopenv.metric.interactive_calc_interruption_001")


def _number(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _find_label_row(ws, label):
    label = label.lower()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None and label in str(value).strip().lower():
            return row
    return None


def check_total_and_average_rows(result_path, expected=None, **options):
    """Verify Total and Average rows contain correct values for Jan-Jun."""
    if result_path is None or not os.path.exists(result_path):
        return 0.0

    expected = expected or {}
    tolerance = float(expected.get("tolerance", 0.01))

    try:
        import openpyxl

        wb = openpyxl.load_workbook(result_path, data_only=True)
        ws = wb.active
        total_row = _find_label_row(ws, "total")
        average_row = _find_label_row(ws, "average")
        if total_row is None or average_row is None or average_row != total_row + 1:
            return 0.0

        data_start = int(expected.get("data_start_row", 2))
        data_end = total_row - 1
        for col in expected.get("value_columns", [2, 3, 4, 5, 6, 7]):
            values = []
            for row in range(data_start, data_end + 1):
                value = _number(ws.cell(row=row, column=col).value)
                if value is None:
                    return 0.0
                values.append(value)

            total_value = _number(ws.cell(row=total_row, column=col).value)
            average_value = _number(ws.cell(row=average_row, column=col).value)
            if total_value is None or average_value is None:
                return 0.0
            if abs(total_value - sum(values)) > tolerance:
                return 0.0
            if abs(average_value - (sum(values) / len(values))) > tolerance:
                return 0.0

        return 1.0
    except Exception as exc:
        logger.warning("check_total_and_average_rows failed: %s", exc)
        return 0.0
    finally:
        try:
            wb.close()
        except Exception:
            pass
