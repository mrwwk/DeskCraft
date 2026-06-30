"""
Evaluator for interactive_calc_error_correction_001:
- Verify spreadsheet data is sorted by March column descending
- Verify Total row exists with correct SUM values for columns B-G
- Verify file was saved (implicitly via reading the .xlsx)
"""
import logging
import os
import traceback

logger = logging.getLogger("desktopenv.metric.calc_sort_and_total")


def check_calc_sort_and_total(result_path, expected=None, **options):
    """
    Check that the LibreOffice Calc spreadsheet has been:
    1. Sorted by the specified column in descending order
    2. Has a Total row with SUM values for each numeric column

    Args:
        result_path: Local path to the .xlsx file pulled from VM
        expected: Dict with rules config (total_row, sort_column, etc.)
        **options: Additional options (not used)

    Returns:
        float: 1.0 if all checks pass, 0.0 otherwise
    """
    if result_path is None or not os.path.exists(result_path):
        logger.warning("Result file not found: %s", result_path)
        return 0.0

    # Default expected rules
    if expected is None:
        expected = {}
    total_row = int(expected.get("total_row", 12))
    sort_column = int(expected.get("sort_column", 4))
    sort_direction = expected.get("sort_direction", "descending")
    sum_columns = expected.get("sum_columns", [2, 3, 4, 5, 6, 7])
    data_start_row = int(expected.get("data_start_row", 2))
    data_end_row = int(expected.get("data_end_row", 11))
    tolerance = float(expected.get("tolerance", 0.01))

    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl is not installed, cannot evaluate .xlsx file")
        return 0.0

    try:
        wb = openpyxl.load_workbook(result_path, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error("Failed to open workbook %s: %s\n%s", result_path, e, traceback.format_exc())
        return 0.0

    try:
        # --- Check 1: Total label in column A at total_row ---
        total_label_cell = ws.cell(row=total_row, column=1).value
        if total_label_cell is None:
            logger.warning("Total label cell A%d is empty", total_row)
            return 0.0
        if "total" not in str(total_label_cell).strip().lower():
            logger.warning("Total label cell A%d does not contain 'total': '%s'", total_row, total_label_cell)
            return 0.0

        # --- Check 2: Sort column values are in descending order ---
        sort_values = []
        for row in range(data_start_row, data_end_row + 1):
            val = ws.cell(row=row, column=sort_column).value
            if val is None:
                logger.warning("Sort column cell (%d, %d) is empty", row, sort_column)
                return 0.0
            try:
                sort_values.append(float(val))
            except (ValueError, TypeError):
                logger.warning("Sort column cell (%d, %d) is not numeric: %s", row, sort_column, val)
                return 0.0

        if sort_direction == "descending":
            for i in range(len(sort_values) - 1):
                if sort_values[i] < sort_values[i + 1]:
                    logger.warning(
                        "Sort order violation at index %d: %s < %s (not descending)",
                        i, sort_values[i], sort_values[i + 1]
                    )
                    return 0.0
        else:
            for i in range(len(sort_values) - 1):
                if sort_values[i] > sort_values[i + 1]:
                    logger.warning(
                        "Sort order violation at index %d: %s > %s (not ascending)",
                        i, sort_values[i], sort_values[i + 1]
                    )
                    return 0.0

        logger.info("Sort check passed: %d values in %s order", len(sort_values), sort_direction)

        # --- Check 3: Total row SUM values for specified columns ---
        for col in sum_columns:
            # Compute expected sum from data rows
            col_sum = 0.0
            for row in range(data_start_row, data_end_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    try:
                        col_sum += float(val)
                    except (ValueError, TypeError):
                        logger.warning("Non-numeric value in data cell (%d, %d): %s", row, col, val)
                        return 0.0

            total_val = ws.cell(row=total_row, column=col).value
            if total_val is None:
                logger.warning("Total cell (%d, %d) is empty", total_row, col)
                return 0.0
            try:
                total_val_float = float(total_val)
            except (ValueError, TypeError):
                logger.warning("Total cell (%d, %d) is not numeric: %s", total_row, col, total_val)
                return 0.0

            if abs(total_val_float - col_sum) > tolerance:
                logger.warning(
                    "SUM mismatch at column %d: expected %.2f, got %.2f (diff=%.4f)",
                    col, col_sum, total_val_float, abs(total_val_float - col_sum)
                )
                return 0.0

            logger.info("Column %d SUM check passed: %.2f", col, total_val_float)

    finally:
        try:
            wb.close()
        except Exception:
            pass

    logger.info("All checks passed for %s", result_path)
    return 1.0
